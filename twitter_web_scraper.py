import requests
from bs4 import BeautifulSoup
import openpyxl

def scrape_twitter_discover(search_text_list):
    base_url = "https://twitter.com/i/discover/flow"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.182 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    }

    all_urls = []
    page = 1
    while True:
        params = {
            "page": page,
            "include_available_features": "1",
            "include_entities": "1",
            "is_prefetch": "0",
            "is_alpha": "0",
            "lang": "en",
            "ref_src": "twsrc^tfw",
        }

        response = requests.get(base_url, params=params, headers=headers)
        if response.status_code != 200:
            print(f"Error occurred while fetching page {page}")
            break

        soup = BeautifulSoup(response.content, "html.parser")
        tweet_links = soup.find_all("a", attrs={"role": "link", "dir": "auto"})
        urls = [link["href"] for link in tweet_links if all(text.lower() in link.text.lower() for text in search_text_list)]
        if not urls:
            break

        # Filter out already saved URLs
        new_urls = [url for url in urls if url not in all_urls]
        all_urls.extend(new_urls)
        page += 1

    return all_urls

def load_saved_urls(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        urls = [cell.value for cell in sheet["A"] if cell.value]
        workbook.close()
        return urls
    except FileNotFoundError:
        return []

def save_urls_to_excel(urls, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for i, url in enumerate(urls):
        sheet.cell(row=i + 1, column=1).value = url

    workbook.save(output_file)
    print(f"URLs saved to {output_file}")

# Prompt the user for input
search_text_input = input("Enter the search text (comma-separated): ")
search_text_list = [text.strip() for text in search_text_input.split(",")]
output_file = input("Enter the output file name: ")

saved_urls = load_saved_urls(output_file)
new_urls = scrape_twitter_discover(search_text_list)
unique_urls = list(set(saved_urls + new_urls))
save_urls_to_excel(unique_urls, output_file)